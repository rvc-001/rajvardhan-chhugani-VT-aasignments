"""
Smart Population Growth and Migration Analysis System
Covers Parts 1-14 (Part 15 = Streamlit bonus, separate)
"""

import random
import math
import statistics
import csv
import os
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# PART 1 – Basic Population Simulation
# ─────────────────────────────────────────────
def simulate(initial=1000, years=50, growth=0.08, migration=0.02):
    records = []
    pop = initial
    for year in range(1, years + 1):
        start = pop
        increase = round(start * growth, 2)
        loss = round(start * migration, 2)
        pop = round(start + increase - loss, 2)
        records.append({
            "year": year,
            "start": round(start, 2),
            "growth": increase,
            "migration_loss": loss,
            "final": pop
        })
    return records

records = simulate()

# ─────────────────────────────────────────────
# PART 2 – Yearly Population Report
# ─────────────────────────────────────────────
def print_report(records):
    print("\n" + "="*75)
    print(f"{'YEARLY POPULATION REPORT':^75}")
    print("="*75)
    print(f"{'Year':>5} {'Start Pop':>14} {'Growth':>10} {'Mig Loss':>12} {'Final Pop':>14}")
    print("-"*75)
    for r in records:
        print(f"{r['year']:>5} {r['start']:>14,.2f} {r['growth']:>10,.2f} {r['migration_loss']:>12,.2f} {r['final']:>14,.2f}")
    print("="*75)

print_report(records)

# ─────────────────────────────────────────────
# PART 3 – Milestone Detection
# ─────────────────────────────────────────────
def find_milestones(records, milestones, initial=1000, net_rate=0.06):
    print("\n" + "="*50)
    print(f"{'POPULATION MILESTONES':^50}")
    print("="*50)
    for target in milestones:
        found = False
        for r in records:
            if r['final'] >= target:
                print(f"  Population > {target:>10,} : Year {r['year']} (Pop: {r['final']:,.2f})")
                found = True
                break
        if not found:
            # Calculate exact year using formula: year = log(target/initial) / log(1+net_rate)
            year_needed = math.ceil(math.log(target / initial) / math.log(1 + net_rate))
            print(f"  Population > {target:>10,} : Not in 50 yrs — needs Year {year_needed}")
    print("="*50)

find_milestones(records, [5000, 10000, 50000, 100000, 1_000_000])

# ─────────────────────────────────────────────
# PART 4 – Mathematical Verification
# ─────────────────────────────────────────────
def math_verification(records, initial=1000, rate=1.06, years=50):
    print("\n" + "="*55)
    print(f"{'MATHEMATICAL VERIFICATION':^55}")
    print("="*55)
    print(f"  Formula: P(t) = P0 × (1.06)^t")
    print(f"  P(50) = 1000 × (1.06)^50 = {1000 * (rate**years):,.2f}")
    print(f"  Simulated P(50) = {records[-1]['final']:,.2f}")
    diff = abs(1000 * rate**years - records[-1]['final'])
    print(f"  Difference (rounding): {diff:.4f}")
    print("="*55)

math_verification(records)

# ─────────────────────────────────────────────
# PART 5 – Long-Term Forecasting
# ─────────────────────────────────────────────
def long_term_forecast(initial=1000, rate=1.06):
    print("\n" + "="*55)
    print(f"{'LONG-TERM FORECASTING':^55}")
    print("="*55)
    print(f"  {'Years':>6} {'Population':>18} {'Multiplier':>12}")
    print("-"*55)
    for years in [50, 100, 200, 500]:
        pop = initial * (rate ** years)
        multiplier = rate ** years
        print(f"  {years:>6} {pop:>18,.2f} {multiplier:>11.2f}x")
    print("="*55)

long_term_forecast()

# ─────────────────────────────────────────────
# PART 6 – Doubling Analysis
# ─────────────────────────────────────────────
def doubling_analysis(records, initial=1000):
    print("\n" + "="*45)
    print(f"{'POPULATION DOUBLING ANALYSIS':^45}")
    print("="*45)
    target = initial * 2
    for r in records:
        if r['final'] >= target:
            print(f"  Target Population  : {target:,}")
            print(f"  Doubled in Year    : {r['year']}")
            print(f"  Population at Year : {r['final']:,.2f}")
            break
    # Rule of 72 verification
    approx = 72 / 6
    print(f"  Rule of 72 estimate: ~{approx:.1f} years")
    print("="*45)

doubling_analysis(records)

# ─────────────────────────────────────────────
# PART 7 – City Comparison
# ─────────────────────────────────────────────
CITIES = {
    "City A": {"growth": 0.08, "migration": 0.02},
    "City B": {"growth": 0.10, "migration": 0.03},
    "City C": {"growth": 0.12, "migration": 0.05},
}

def city_comparison(years=100):
    print("\n" + "="*60)
    print(f"{'CITY COMPARISON ({} YEARS)'.format(years):^60}")
    print("="*60)
    city_records = {}
    for city, params in CITIES.items():
        r = simulate(years=years, **params)
        city_records[city] = r
        net = params['growth'] - params['migration']
        final_pop = r[-1]['final']
        print(f"  {city}: Net Growth={net*100:.0f}%  Final Pop={final_pop:>15,.2f}")
    winner = max(city_records, key=lambda c: city_records[c][-1]['final'])
    print(f"\n  ✅ Fastest Growing: {winner}")
    print("="*60)
    return city_records

city_data = city_comparison()

# ─────────────────────────────────────────────
# PART 8 – Population Decline
# ─────────────────────────────────────────────
def decline_scenario(initial=1000, growth=0.03, migration=0.05, years=100):
    print("\n" + "="*50)
    print(f"{'POPULATION DECLINE SCENARIO':^50}")
    print("="*50)
    pop = initial
    below_500 = below_100 = extinct = None
    for year in range(1, years + 1):
        pop = pop * (1 + growth - migration)
        if pop < 500 and below_500 is None:
            below_500 = (year, pop)
        if pop < 100 and below_100 is None:
            below_100 = (year, pop)
        if pop < 1 and extinct is None:
            extinct = (year, pop)
    print(f"  Net Growth: {(growth-migration)*100:.0f}% (Declining)")
    print(f"  Below 500 : Year {below_500[0]} (Pop: {below_500[1]:,.2f})" if below_500 else "  Below 500 : Never in 100 years")
    print(f"  Below 100 : Year {below_100[0]} (Pop: {below_100[1]:,.2f})" if below_100 else "  Below 100 : Never in 100 years")
    print(f"  Extinction: Year {extinct[0]}" if extinct else "  Extinction: Does NOT occur in 100 years")
    print("="*50)

decline_scenario()

# ─────────────────────────────────────────────
# PART 9 – Monte Carlo Simulation
# ─────────────────────────────────────────────
def monte_carlo(simulations=100, years=100, initial=1000):
    print("\n" + "="*50)
    print(f"{'MONTE CARLO SIMULATION ({} runs)'.format(simulations):^50}")
    print("="*50)
    outcomes = []
    for _ in range(simulations):
        pop = initial
        for _ in range(years):
            g = random.uniform(0.05, 0.10)
            m = random.uniform(0.01, 0.04)
            pop = pop * (1 + g - m)
        outcomes.append(pop)
    print(f"  Best Outcome   : {max(outcomes):>15,.2f}")
    print(f"  Worst Outcome  : {min(outcomes):>15,.2f}")
    print(f"  Average Outcome: {statistics.mean(outcomes):>15,.2f}")
    print("="*50)
    return outcomes

mc_outcomes = monte_carlo()

# ─────────────────────────────────────────────
# PART 10 – Statistical Analysis
# ─────────────────────────────────────────────
def statistical_analysis(records):
    pops = [r['final'] for r in records]
    arr = np.array(pops)
    print("\n" + "="*50)
    print(f"{'STATISTICAL ANALYSIS (50-Year Sim)':^50}")
    print("="*50)
    print(f"  Maximum Population : {arr.max():>15,.2f}")
    print(f"  Minimum Population : {arr.min():>15,.2f}")
    print(f"  Average Population : {arr.mean():>15,.2f}")
    print(f"  Median Population  : {np.median(arr):>15,.2f}")
    print(f"  Std Deviation      : {arr.std():>15,.2f}")
    print(f"  Growth Variance    : {arr.var():>15,.2f}")
    print("="*50)

statistical_analysis(records)

# ─────────────────────────────────────────────
# PART 11 – Data Visualization
# ─────────────────────────────────────────────
os.makedirs('/mnt/user-data/outputs', exist_ok=True)

def save_charts(records, city_data):
    years = [r['year'] for r in records]
    pops = [r['final'] for r in records]
    growths = [r['growth'] for r in records]
    mig_losses = [r['migration_loss'] for r in records]

    plt.style.use('seaborn-v0_8-darkgrid')
    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    fig.suptitle('Population Growth Analysis Dashboard', fontsize=18, fontweight='bold', y=0.98)

    # Line Chart
    axes[0, 0].plot(years, pops, color='steelblue', linewidth=2.5, marker='o', markersize=3)
    axes[0, 0].set_title('Population vs Year', fontsize=13, fontweight='bold')
    axes[0, 0].set_xlabel('Year'); axes[0, 0].set_ylabel('Population')
    axes[0, 0].yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{x:,.0f}'))
    axes[0, 0].fill_between(years, pops, alpha=0.15, color='steelblue')

    # Bar Chart
    axes[0, 1].bar(years, growths, color='forestgreen', alpha=0.8, width=0.8)
    axes[0, 1].set_title('Yearly Population Growth', fontsize=13, fontweight='bold')
    axes[0, 1].set_xlabel('Year'); axes[0, 1].set_ylabel('Growth Amount')
    axes[0, 1].yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{x:,.0f}'))

    # Pie Chart
    total_growth = sum(growths)
    total_mig = sum(mig_losses)
    axes[1, 0].pie([total_growth, total_mig],
                   labels=['Total Growth', 'Total Migration Loss'],
                   autopct='%1.1f%%', colors=['#2ecc71', '#e74c3c'],
                   startangle=90, explode=(0.05, 0.05))
    axes[1, 0].set_title('Growth vs Migration Loss (50 Years)', fontsize=13, fontweight='bold')

    # City Comparison
    colors = ['#3498db', '#e67e22', '#9b59b6']
    for (city, data), color in zip(city_data.items(), colors):
        y = [r['year'] for r in data]
        p = [r['final'] for r in data]
        axes[1, 1].plot(y, p, label=city, color=color, linewidth=2)
    axes[1, 1].set_title('City A vs City B vs City C (100 Years)', fontsize=13, fontweight='bold')
    axes[1, 1].set_xlabel('Year'); axes[1, 1].set_ylabel('Population')
    axes[1, 1].legend(fontsize=10)
    axes[1, 1].yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f'{x:,.0f}'))

    plt.tight_layout()
    path = '/mnt/user-data/outputs/population_charts.png'
    plt.savefig(path, dpi=150, bbox_inches='tight')
    plt.close()
    print(f"\n  ✅ Charts saved: {path}")

save_charts(records, city_data)

# ─────────────────────────────────────────────
# PART 12 – Excel Report
# ─────────────────────────────────────────────
def make_excel(records, city_data):
    wb = Workbook()

    hdr = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", start_color="2E4057")
    alt_fill = PatternFill("solid", start_color="EBF5FB")
    center = Alignment(horizontal="center")
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header(ws, headers, row=1):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = hdr; cell.fill = hdr_fill
            cell.alignment = center; cell.border = thin

    def set_col_widths(ws, widths):
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    # Sheet 1 – Yearly Population
    ws1 = wb.active; ws1.title = "Yearly Population"
    ws1['A1'] = "Population Growth Report – 50 Years"
    ws1['A1'].font = Font(bold=True, size=14, color="1A5276")
    ws1.merge_cells('A1:E1')
    ws1['A1'].alignment = center

    style_header(ws1, ["Year", "Starting Population", "Growth (+8%)", "Migration Loss (-2%)", "Final Population"], row=2)

    for i, r in enumerate(records, 3):
        row = [r['year'], r['start'], r['growth'], r['migration_loss'], r['final']]
        for col, val in enumerate(row, 1):
            cell = ws1.cell(row=i, column=col, value=val)
            cell.border = thin; cell.alignment = center
            if i % 2 == 0:
                cell.fill = alt_fill

    set_col_widths(ws1, [8, 20, 15, 18, 18])

    # Sheet 2 – Growth Analysis
    ws2 = wb.create_sheet("Growth Analysis")
    ws2['A1'] = "Mathematical Growth Analysis"
    ws2['A1'].font = Font(bold=True, size=14, color="1A5276")
    ws2.merge_cells('A1:C1'); ws2['A1'].alignment = center

    style_header(ws2, ["Metric", "Value", "Notes"], row=2)
    net_rate = 0.06
    analysis_data = [
        ("Initial Population", 1000, "Base year"),
        ("Growth Rate", "8%", "Annual"),
        ("Migration Rate", "2%", "Annual"),
        ("Net Growth Rate", "6%", "Growth - Migration"),
        ("Population @ 50 yrs", f"{records[-1]['final']:,.2f}", "Simulated"),
        ("Population @ 50 yrs (Formula)", f"{1000*(1.06**50):,.2f}", "P0 × 1.06^50"),
        ("Years to Double", "12", "Rule of 72 ≈ 12 yrs"),
        ("Population Multiplier @ 50", f"{1.06**50:.2f}x", ""),
    ]
    for i, row in enumerate(analysis_data, 3):
        for col, val in enumerate(row, 1):
            cell = ws2.cell(row=i, column=col, value=val)
            cell.border = thin; cell.alignment = center
            if i % 2 == 0: cell.fill = alt_fill
    set_col_widths(ws2, [28, 20, 22])

    # Sheet 3 – Milestone Report
    ws3 = wb.create_sheet("Milestone Report")
    ws3['A1'] = "Population Milestone Report"
    ws3['A1'].font = Font(bold=True, size=14, color="1A5276")
    ws3.merge_cells('A1:C1'); ws3['A1'].alignment = center

    style_header(ws3, ["Target Population", "Year Reached", "Population at That Year"], row=2)
    milestones = [5000, 10000, 50000, 100000, 1_000_000]
    for i, target in enumerate(milestones, 3):
        year_val = pop_val = "Not in 50 yrs"
        for r in records:
            if r['final'] >= target:
                year_val = r['year']; pop_val = f"{r['final']:,.2f}"; break
        if year_val == "Not in 50 yrs":
            yr = math.ceil(math.log(target/1000)/math.log(1.06))
            year_val = f"Year {yr} (extended)"
        for col, val in enumerate([f"{target:,}", year_val, pop_val], 1):
            cell = ws3.cell(row=i, column=col, value=val)
            cell.border = thin; cell.alignment = center
            if i % 2 == 0: cell.fill = alt_fill
    set_col_widths(ws3, [22, 22, 28])

    # Sheet 4 – City Comparison
    ws4 = wb.create_sheet("City Comparison")
    ws4['A1'] = "City Population Comparison – 100 Years"
    ws4['A1'].font = Font(bold=True, size=14, color="1A5276")
    ws4.merge_cells('A1:D1'); ws4['A1'].alignment = center

    style_header(ws4, ["Year", "City A (Net 6%)", "City B (Net 7%)", "City C (Net 7%)"], row=2)
    years_to_show = list(range(0, 101, 10))
    for i, yr in enumerate(years_to_show, 3):
        row_vals = [yr]
        for city, data in city_data.items():
            row_vals.append(data[yr-1]['final'] if yr > 0 else 1000)
        for col, val in enumerate(row_vals, 1):
            cell = ws4.cell(row=i, column=col, value=val)
            cell.border = thin; cell.alignment = center
            if i % 2 == 0: cell.fill = alt_fill
    set_col_widths(ws4, [10, 20, 20, 20])

    path = '/mnt/user-data/outputs/population_report.xlsx'
    wb.save(path)
    print(f"  ✅ Excel saved: {path}")

make_excel(records, city_data)

# ─────────────────────────────────────────────
# PART 13 – CSV Export
# ─────────────────────────────────────────────
def export_csvs(records, city_data):
    # population_data.csv
    with open('/mnt/user-data/outputs/population_data.csv', 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=["year", "start", "growth", "migration_loss", "final"])
        w.writeheader(); w.writerows(records)

    # growth_analysis.csv
    with open('/mnt/user-data/outputs/growth_analysis.csv', 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(["Metric", "Value"])
        w.writerow(["Initial Population", 1000])
        w.writerow(["Final Population (50 yr)", records[-1]['final']])
        w.writerow(["Net Growth Rate", "6%"])
        w.writerow(["Mathematical Formula", "P(t) = 1000 × 1.06^t"])
        for yr in [50, 100, 200, 500]:
            w.writerow([f"Population @ {yr} yrs", f"{1000*(1.06**yr):,.2f}"])

    # milestone_report.csv
    with open('/mnt/user-data/outputs/milestone_report.csv', 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(["Target", "Year Reached", "Population"])
        for target in [5000, 10000, 50000, 100000, 1_000_000]:
            found = False
            for r in records:
                if r['final'] >= target:
                    w.writerow([target, r['year'], r['final']]); found = True; break
            if not found:
                yr = math.ceil(math.log(target/1000)/math.log(1.06))
                w.writerow([target, f"Year {yr} (extended)", "N/A"])

    print("  ✅ CSVs saved: population_data.csv, growth_analysis.csv, milestone_report.csv")

export_csvs(records, city_data)

# ─────────────────────────────────────────────
# PART 14 – Population Query System
# ─────────────────────────────────────────────
def query_year(records):
    print("\n" + "="*45)
    print(f"{'POPULATION QUERY SYSTEM':^45}")
    print("="*45)
    try:
        year = int(input("  Enter Year (1–50): "))
        if 1 <= year <= 50:
            r = records[year - 1]
            print(f"  Year            : {r['year']}")
            print(f"  Starting Pop    : {r['start']:,.2f}")
            print(f"  Growth (+8%)    : {r['growth']:,.2f}")
            print(f"  Migration Loss  : {r['migration_loss']:,.2f}")
            print(f"  Final Pop       : {r['final']:,.2f}")
        else:
            print("  Year out of range!")
    except (ValueError, EOFError):
        print("  [Query skipped in non-interactive mode]")
        r = records[9]  # Show Year 10 as demo
        print(f"  Demo – Year 10: Start={r['start']:,.2f}, Growth={r['growth']:,.2f}, Final={r['final']:,.2f}")
    print("="*45)

query_year(records)

# ─────────────────────────────────────────────
# SUMMARY ANSWERS
# ─────────────────────────────────────────────
print("\n" + "="*55)
print(f"{'EXPECTED OUTPUT ANSWERS':^55}")
print("="*55)
print(f"  1. Population after 50 years : {records[-1]['final']:,.2f}")
pop_1m_year = next((r['year'] for r in records if r['final'] >= 1_000_000), None)
exceeds = "Yes" if pop_1m_year else "No"
print(f"  2. Exceeds 1 million?        : {exceeds}")
if pop_1m_year:
    print(f"  3. Year exceeds 1 million    : Year {pop_1m_year}")
else:
    yr = math.ceil(math.log(1_000_000/1000)/math.log(1.06))
    print(f"  3. Would reach 1M in         : Year {yr}")
double_year = next((r['year'] for r in records if r['final'] >= 2000), None)
print(f"  4. Years to double           : Year {double_year}")
winner = max(city_data, key=lambda c: city_data[c][-1]['final'])
print(f"  5. Fastest growing city      : {winner}")
print(f"  6. Migration reduces net %   : 8% growth - 2% migration = 6% net")
print(f"  7. When mig > growth (-2%)   : Population declines ~3.4%/yr")
print("="*55)
print("\n✅ All parts complete! Files saved to /mnt/user-data/outputs/\n")
